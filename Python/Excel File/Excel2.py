#READ ONE CELL DATA

#import package
import openpyxl
wk=openpyxl.load_workbook('D:\\Test.xlsx')

# print(wk.Sheet1(1,3))

sh=wk['Sheet1']  #sheet object
print(sh['A3'].value)
print(sh['B4'].value)

sk=wk['Hello']
print(sk['A1'].value)

c1=sh.cell(3,1)   #row,columnn by default
print(c1.value)

c2=sh.cell(column=3,row=2)     #by creating cell objects
print(c2.value)

print(c1.row)
print(c2.row)
