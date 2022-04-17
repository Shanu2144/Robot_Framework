#READ ALL ROWS & CELL DATA(COMPLETE DATA)
#import package
import openpyxl

wk=openpyxl.load_workbook('D:\\Test.xlsx')
sh=wk['Sheet1']

#find rows having data
rows=sh.max_row
columns=sh.max_column

print('total rows are ' + str(rows))
print('total columns are ' + str(columns))

for i in range(1,rows+1):
    for j in range(1,columns+1):
        c=sh.cell(i,j)
        print(c.value)

#OR

for r in sh['A1':'C4']:
    for c in r:
        print(c.value)

