# Import Package
import openpyxl

def fetch_number_of_rows(Sheetname,filepath):
    wk = openpyxl.load_workbook(filepath)
    sh = wk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell,filepath):
    wk = openpyxl.load_workbook(filepath)
    sh = wk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value